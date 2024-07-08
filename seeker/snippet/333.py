#date: 2024-07-08T17:05:41Z
#url: https://api.github.com/gists/a7f60a8a8e8a469dacfbdf998d53aa7f
#owner: https://api.github.com/users/zero-ljz

import sys, os, datetime
import sqlite3
from openpyxl import load_workbook

# Excel文件存放目录
dir_path = r'.\'

conn = sqlite3.connect(r'.\db.sqlite3')
cursor = conn.cursor()
cursor.row_factory = sqlite3.Row # 设置返回字典

def handle_data_row(row):
    # print(row) # 打印出当前将要导入的行


    cursor.execute("SELECT * FROM Customer WHERE name = ?", (row['合作客户'],))
    customer = cursor.fetchone()
    if not customer:
        print('找不到客户:', row['合作客户'], '，请检查系统中的客户表')

    # 查询出对应用户id
    cursor.execute("SELECT * FROM auth_user WHERE first_name = ?", (row['账号编辑'],))
    author = cursor.fetchone()
    if not author:
        print('找不到账号编辑（用户）:', row['账号编辑'])

    DateObj = datetime.datetime(year=1899, month=12, day=30) 
    DateObj += datetime.timedelta(days=row['时间']) # row['时间']会返回小数，通过这个转成日期
    r_date = DateObj.strftime("%Y-%m-%d") # 记得在setting中将 USE_TZ设为False
    # r_date = row['时间'].strftime("%Y-%m-%d")
    # 根据账号、日期、文章位置查找相应记录
    cursor.execute(
        "SELECT * FROM Record WHERE account_id = ? AND date = ? AND post_location = ?", 
        (account['id'], r_date, int(row['文章位置']))
    )
    record = cursor.fetchone()
    if not record:
        print(f"账号{account['id']}、日期{r_date}、文章位置{r_date}，没有对应记录，如果是月初请先执行数据生成")

    if hz_mode is None or not customer or not account or not author or not record:
        print('警告: 由于遇到问题，此行未导入')
        return

    r = dict(record)
    # print(r) # 打印找到的记录

    # 修改找到的这条记录
    r['author_id'] = author['id']
    r['fans_count'] = row['粉丝量']

    r['hz_mode'] = hz_mode

    filter = {"id": r["id"]}
    r.pop('id')
    query = f"UPDATE Record SET {', '.join([f'{key} = ?' for key in r])} WHERE {' AND '.join([f'{key} = ?' for key in filter])}"
    args = tuple(r.values()) + tuple(filter.values())
    print(args)
    cursor.execute(query, args)
    conn.commit()


file_list = os.listdir(dir_path)
print('开始遍历表格行')
for file_name in file_list:
    wb = load_workbook(filename = dir_path + '\\' + file_name)

    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
    else:
        ws = wb.active

    # print(ws['D18'].value)

    first_row_values = tuple(cell.value for cell in ws[1])
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False)): # 从第二行开始读取
        row_num = row[0].row
        row_values = tuple(cell.value for cell in row)
        # print('i:', i, 'row_num:', row_num, 'row_values:', row_values)
        if row_values[0] is None: # 第一列为空时跳过
            continue

        print('当前导入Excel第', row_num, '行')
        row_dict = {k: v for k, v in zip(first_row_values, row_values)} # 将首行（表头行）和当前行合并为字典
        handle_data_row(row_dict)
        # if row_num == 5:
        #     break

