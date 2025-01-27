#date: 2025-01-27T16:44:44Z
#url: https://api.github.com/gists/11ac1d7a482ed754687e2359640b48c5
#owner: https://api.github.com/users/adgedenkers

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def modify_excel(input_path, output_path):
    # Load the workbook and select the relevant sheet
    workbook = load_workbook(input_path)
    sheet1 = workbook['Sheet1']

    # Step 1: Freeze header row and columns A-F
    sheet1.freeze_panes = 'G2'

    # Step 2: Make header row text bold
    bold_font = Font(bold=True)
    for cell in sheet1[1]:
        cell.font = bold_font

    # Step 3: Apply autofilter to all columns
    sheet1.auto_filter.ref = sheet1.dimensions

    # Step 4: Apply green fill to headings in columns A-F
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    for col in range(1, 7):  # Columns A-F (1 to 6)
        sheet1.cell(row=1, column=col).fill = green_fill

    # Save the modified workbook to the specified output path
    workbook.save(output_path)

# Paths for input and output Excel files
input_path = 'dev_sheet.xlsx'  # Replace with your input file path
output_path = 'modified_dev_sheet.xlsx'  # Replace with your desired output file path

# Execute the modification
modify_excel(input_path, output_path)

print(f"Modified Excel file saved to: {output_path}")
