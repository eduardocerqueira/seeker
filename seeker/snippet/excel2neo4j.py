#date: 2022-12-13T16:49:17Z
#url: https://api.github.com/gists/a00132a1d9818ec85fa61b13e9def8bd
#owner: https://api.github.com/users/sorphwer

from typing import Dict
from py2neo import Graph, Node, Relationship
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import traceback
import json
'''
INPUT: sheet, row index
OUTPUT: list
Example : get_row_props(wb_obj.active,1)
'''
def get_row_props(sheet:openpyxl.worksheet.worksheet.Worksheet,row_index:int)->list:
    row_props = []
    for i in range(1,sheet.max_column):
        row_props.append(sheet[get_column_letter(i)+str(row_index)].value)
    return row_props

def get_node_props(header:list,
                   row_props:list,
                   node_scheme:list,
                   hashmap = {},
                  )->Dict[str,list]:
    res = {}
        
    def _get_dict_node_props(header:list,props:list)->dict:#genarate dict format for py2neo
        if len(header) != len(props):
            print('Error: length of header and props is not same in _get_dict_node_props')
            return ''
        else:
            res = dict()
            for i,j in zip(header,props):
                i = i.split('\n')[-1]
                if type(i) == type('str'): i = i.replace(' ','_') 
                i = i.replace('-','_')
                
                if j == None: 
                    res[i] = 'None'
                else:
                    if type(j) == type('str'): 
                        # j = j.replace(' ','_')
                        j = j.strip(' ')
                        if 'j' in hashmap:
                            j = hashmap[j]
                    res[i] = j
            return res
        
    def _no_split_props_gen(node_props_index:list,row_props=row_props)->list:
        _node_header = []
        _node_props = []
        for i in node_props_index:
            _node_header.append(header[i-1])
            _node_props.append(row_props[i-1])
        if len(_node_props)==1:
            if _node_props[0] == None:
                return []
        # return [ _get_neo4j_node_props(_node_header,_node_props)]
        return [ _get_dict_node_props(_node_header,_node_props)]
    
    def _split_props_gen(node_props_index:list,split_col_number=-1,separator=[',','，'])->list:
        if split_col_number == -1 : split_col_number=node_props_index[0]
        split_target_value = row_props[split_col_number-1]
        
        if type(split_target_value)==type('str'):
            split_target_value = split_target_value.replace('\n',',')
            split_target_value = split_target_value.replace('，',',')
            
            for s in separator:
                split_target_value = split_target_value.replace(s,',')
            if ',' in split_target_value:
                split_list =  split_target_value.split(',')
            else:
                return _no_split_props_gen(node_props_index)
            res = []
            for i in split_list:
                if i == ' ':
                    continue
                temp = row_props.copy()
                temp[split_col_number-1] = i
                res += _no_split_props_gen(node_props_index,row_props=temp)
            return res
        else:
            return _no_split_props_gen(node_props_index)
    #Generate custom props for create nodes:
    
    for i in node_scheme:
        if 'split' in i:
            if 'separator' in i:
                res[i['name']] = _split_props_gen(i['col'],i['split'],i['separator'])
            else:
                res[i['name']] = _split_props_gen(i['col'],i['split'])
        else:
            res[i['name']] = _no_split_props_gen(i['col'])
    return res
    
def create_meta(
                file_path:str,
                sheet_name:str,
                header_index:int,
                row_index:int,
                node_schema:list,
                relationship_schema:list
               ):
    nodes_props =  get_node_props(get_row_props(wb_obj[sheet_name],header_index),
                                  get_row_props(wb_obj[sheet_name],row_index),
                                  node_schema)
    
    #init node
    node_obj_list = []

    for node_tag in nodes_props.keys():
        for node in nodes_props[node_tag]:
            node_obj = Node(node_tag,**node)
            node_obj_list.append(node_obj)
            del node_obj
            
    #init relationship
    relationship_obj_list = []
    for node_obj in node_obj_list:
        for schema in relationship_schema:
            if node_obj.has_label(schema['start']):
                #ok, this node should be the start of a relationship, let's find it out
                for node_obj_2 in node_obj_list:
                    if node_obj_2.has_label(schema['end']):
                        relationship_obj = Relationship(node_obj,schema['name'],node_obj_2)
                        relationship_obj_list.append(relationship_obj)
            
            
    return node_obj_list,relationship_obj_list