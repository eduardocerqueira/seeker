#date: 2026-01-23T17:00:45Z
#url: https://api.github.com/gists/6fb54198872d9a0b35141d88a1fa41a1
#owner: https://api.github.com/users/brett6320

#!/usr/bin/env python3
"""
Convert CSV file to XLSX with formatting:
- Bold header row
- Frozen top row
- Auto-filter enabled
- Column widths sized to content (max 128 characters)

Usage:
    python csv_to_xlsx.py /path/to/file.csv
"""

import sys
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font


def convert_csv_to_xlsx(csv_path):
    """Convert CSV to XLSX with formatting applied."""
    
    csv_path = Path(csv_path).resolve()
    
    # Validate input file exists
    if not csv_path.exists():
        print(f"Error: File not found: {csv_path}", file=sys.stderr)
        sys.exit(1)
    
    if csv_path.suffix.lower() != '.csv':
        print(f"Error: File must be a CSV file: {csv_path}", file=sys.stderr)
        sys.exit(1)
    
    # Read CSV data
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
    except Exception as e:
        print(f"Error reading CSV: {e}", file=sys.stderr)
        sys.exit(1)
    
    if not rows:
        print("Error: CSV file is empty", file=sys.stderr)
        sys.exit(1)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    
    # Write data and calculate column widths
    column_widths = {}
    
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Make header row bold
            if row_idx == 1:
                cell.font = Font(bold=True)
            
            # Track max column width (max 128 chars)
            col_letter = cell.column_letter
            value_len = len(str(value)) if value else 0
            max_width = min(value_len + 1, 128)
            
            if col_letter not in column_widths:
                column_widths[col_letter] = max_width
            else:
                column_widths[col_letter] = max(column_widths[col_letter], max_width)
    
    # Apply column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Add autofilter to header row
    if rows:  # Only add filter if there's data
        header_range = f'A1:{ws.cell(row=1, column=len(rows[0])).column_letter}1'
        ws.auto_filter.ref = header_range
    
    # Write to output file
    output_path = csv_path.with_suffix('.xlsx')
    
    try:
        wb.save(output_path)
        print(f"✓ Converted successfully!")
        print(f"  Input:  {csv_path}")
        print(f"  Output: {output_path}")
    except Exception as e:
        print(f"Error writing XLSX: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage: python csv_to_xlsx.py <csv_file_path>", file=sys.stderr)
        sys.exit(1)
    
    convert_csv_to_xlsx(sys.argv[1])
