#date: 2022-01-27T16:59:17Z
#url: https://api.github.com/gists/a6e30db9660646767f3589989664fade
#owner: https://api.github.com/users/Qeswer

import openpyxl
import re
import matplotlib.pyplot as plt
import numpy as np

exel_file = openpyxl.load_workbook("value.xlsx")
book2 = exel_file['разница']


def y_list(x, col):  # функция вытаскивания значений в столбик
    y_list_ = []
    for i in range(1, x.max_row):
        y_list_.append(x.cell(row=i, column=col).value)
    # print(y_list_)
    return y_list_[1:len(y_list_)]


def reg(x):  # функция регулярного выражения
    list1 = []
    list2 = []
    for i in range(len(x)):
        list1.append(re.sub(r'2021-11-29T12:', '', x[i]))
        list2.append(re.findall(r'\d{2}:\d{2}', list1[i]))
        list1[i] = list2[i][0]
    return list1


def sens():
    list1 = []
    for i in range(2, book2.max_column):
        list1.append(y_list(book2, i))
    for i in list1:
        i[0] = float(0)
    return list1


def plot_sens(T: list, S: list) -> list:
    for i in range(len(S)):
        plt.plot(T, S[i])
    plt.show()


def __print__(x):
    for i in x:
        print()
        for j in i:
            print(j, end=" ")


time = np.array(reg(y_list(book2, 1)))
# sensors = sens()
# print(type(sensors[5][9]))
sensors = np.array(sens())
# __print__(sensors)
plot_sens(time, sensors)
