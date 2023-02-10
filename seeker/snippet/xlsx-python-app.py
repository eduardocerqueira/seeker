#date: 2023-02-10T16:50:58Z
#url: https://api.github.com/gists/2de0a34ba19e3cc8591ddc256eefc842
#owner: https://api.github.com/users/hoochiecoo

import re
from openpyxl import load_workbook


def xlsx_to_dict(filename):
    """Converts an XLSX file to a dictionary."""
    # Load the workbook
    wb = load_workbook(filename)

    # Create an empty dictionary
    data = {}

    # Iterate over each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Iterate over each cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                # Add the cell's value to the dictionary using its location as the key
                key = f"{sheet_name}!{cell.column_letter}{cell.row}"
                data[key] = cell.value

    return data


# Convert the XLSX file to a dictionary
data = xlsx_to_dict('to_yaml.xlsx')

print(data)


def replace_cell_references(formula, data):
    """Replaces cell references in an Excel formula with their values from a dictionary."""
    # Remove the first '=' character from the formula
    formula = formula[1:]

    # Split the formula into separate components
    components = re.split(r'([\+\-\*\/\(\)\^])', formula)

    # Replace cell references with their values
    for i, component in enumerate(components):
        if not re.match(r'([\+\-\*\/\(\)\^])', component):
            # Get the cell reference
            cell_reference = component

            # Replace the cell reference with its value
            components[i] = str(data.get(cell_reference, cell_reference))
            #components[i] = "**" if component == "^" else component

        if component == "^":
            components[i] = "**"

    # Join the components back together and return the result
    return (''.join(components))


# Get the formula from the data dictionary
formula = data['REPORT!B1']


def get_sheets_from_file(file_path):
    wb = load_workbook(file_path)
    sheets = wb.sheetnames
    return sheets


def evaluate_formula(file_path, formula):
    sheets = get_sheets_from_file(file_path)

    while any(sheet_name in formula for sheet_name in sheets):
        result = replace_cell_references(formula, data)
        formula = result

    return result


result = evaluate_formula("to_yaml.xlsx", formula)
result_eval = eval(result)
print(result_eval)
