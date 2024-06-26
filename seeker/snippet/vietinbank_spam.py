#date: 2024-06-26T16:28:37Z
#url: https://api.github.com/gists/425a9cf5ce8b0919f6a0a96194086c13
#owner: https://api.github.com/users/loozzi

import time

from openpyxl import load_workbook
from selenium.webdriver import Chrome

driver = Chrome()
url = "https://dangkymuavang.vietinbankgold.vn/"


def fillData(
    driver,
    fullName,
    typeOfId,
    idNumber,
    dateOfBirth,
    sex,
    address,
    dateRange,
    addressRange,
    phoneNumber,
    email,
    amount,
    addressReceive,
):
    driver.find_element(by="name", value="hvt").send_keys(fullName)
    driver.find_element(by="name", value="l").send_keys(typeOfId)
    driver.find_element(by="name", value="s").send_keys(idNumber)
    driver.find_element(by="name", value="ns").send_keys(dateOfBirth)
    driver.find_element(by="name", value="gt").send_keys(sex)
    driver.find_element(by="name", value="dctt").send_keys(address)
    driver.find_element(by="name", value="nc").send_keys(dateRange)
    driver.find_element(by="name", value="dcc").send_keys(addressRange)
    driver.find_element(by="name", value="sdt").send_keys(phoneNumber)
    driver.find_element(by="name", value="ht").send_keys(email)
    driver.find_element(by="name", value="sl").clear()
    driver.find_element(by="name", value="sl").send_keys(amount)
    driver.find_element(by="name", value="dc").send_keys(addressReceive)


workbook = load_workbook(filename="data.xlsx")
sheet = workbook["Sheet1"]

data = []
for row in sheet.iter_rows(
    min_row=1,
    max_row=sheet.max_row,
    min_col=1,
    max_col=sheet.max_column,
    values_only=True,
):
    data.append(row)

workbook.close()

for row in data[1:]:
    driver.get(url)
    fillData(driver, *row)

    input("Press Enter to continue...")
    time.sleep(1)
